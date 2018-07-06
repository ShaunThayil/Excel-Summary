import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.worksheet.dimensions import ColumnDimension


def unique_index(sheet):    #Returns a list of unique indexes 
    list1=list(sheet.index.values)
    list2=[]
    for i in list1:
        if i not in list2:
         list2.append(i)
    return list2

def col_names(sheet):   #Returns The Column Header Names Without Unwanted Columns
    col_name=list(sheet.columns.values)
    col_name.remove("DATE")
    col_name.remove("TIME")
    return col_name

def create_sheets(wb,col_names,savename): #Create Sheets For Each Column In Target File And Remove The Initial Sheet
    for i in col_names:
        wb.create_sheet(title=i)
    list_names=wb.get_sheet_names()
    rm_sheet=wb.get_sheet_by_name(name=list_names[0])
    wb.remove_sheet(rm_sheet)
    wb.save(savename)

def daily_means_max_min(new_workbook,sheet,index_list,col_name,target_sheet,target_filename): #To Add Values To The New Sheet
  index_list_name=[]
  for i in index_list:      #To Convert Index datetime value To Date Only
      val1=str(i)
      val=val1[:10]
      index_list_name.append(val)    
  
  values=[]
  for i in index_list:
     mean=sheet.loc[i,col_name].mean()
     maximum=sheet.loc[i,col_name].max()
     minimum=sheet.loc[i,col_name].min()
     values.append([mean,maximum,minimum])
  df=pd.DataFrame(values,index=index_list_name,columns=["Avg","Max","Min"])  #Add Value To DataFrame

  for r in dataframe_to_rows(df, index=True, header=True):  #To Convert DataFrame To Rows
    target_sheet.append(r)

  for cell in target_sheet['A'] + target_sheet[1]:
    cell.style = 'Pandas'
  target_sheet.cell(row=1,column=1).value="DATE" #To Assign Header To First Column
  new_workbook.save(target_filename)             #To Save To Target File
  
     